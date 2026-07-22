"""The booking-history workbook: overview, customers, every request."""
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from bot.locales import t
from bot.timeutil import fmt_minutes

_HEADER = Font(bold=True)


def _write_header(sheet, titles: list[str]) -> None:
    sheet.append(titles)
    for cell in sheet[1]:
        cell.font = _HEADER


def _autosize(sheet) -> None:
    for column in sheet.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = min(width + 2, 60)


def _when(booking) -> str:
    """The parsed slot, or the raw text when it was never resolved."""
    if booking.date is None or booking.start_minute is None:
        return booking.when_text
    end = booking.start_minute + booking.duration_hours * 60
    return f"{booking.date.isoformat()} {fmt_minutes(booking.start_minute)}–{fmt_minutes(end)}"


def build_stats_workbook(overview, users, bookings, today: date, lang: str) -> bytes:
    book = Workbook()

    summary = book.active
    summary.title = t("xls_sheet_overview", lang)
    summary.append([t("xls_title", lang, date=today.isoformat())])
    summary["A1"].font = _HEADER
    summary.append([])
    for key, value in (
        ("xls_metric_users", overview["users"]),
        ("xls_metric_requests", overview["requests"]),
        ("xls_metric_accepted", overview["accepted"]),
        ("xls_metric_rejected", overview["rejected"]),
        ("xls_metric_pending", overview["pending"]),
        ("xls_metric_today", overview["today"]),
        ("xls_metric_people", overview["people"]),
        ("xls_metric_hours", overview["hours"]),
    ):
        summary.append([t(key, lang), value])
    _autosize(summary)

    customers = book.create_sheet(t("xls_sheet_customers", lang))
    _write_header(customers, [
        t("xls_h_name", lang), t("xls_h_phone", lang), t("xls_h_language", lang),
        t("xls_h_requests", lang), t("xls_h_accepted", lang), t("xls_h_people", lang),
        t("xls_h_first", lang), t("xls_h_last", lang),
    ])
    for u in users:
        customers.append([
            u.name, u.phone, u.language, u.requests, u.accepted, u.people,
            u.first_seen.isoformat() if u.first_seen else "",
            u.last_booking.isoformat() if u.last_booking else "",
        ])
    _autosize(customers)

    history = book.create_sheet(t("xls_sheet_bookings", lang))
    _write_header(history, [
        t("xls_h_date", lang), t("xls_h_when", lang), t("xls_h_requested", lang),
        t("xls_h_people", lang), t("xls_h_hours", lang), t("xls_h_status", lang),
        t("xls_h_customer", lang), t("xls_h_phone", lang), t("xls_h_created", lang),
    ])
    for b in bookings:
        history.append([
            b.date.isoformat() if b.date else "",
            _when(b),
            b.when_text,
            b.people_count,
            b.duration_hours,
            t(f"status_{b.status}", lang),
            b.full_name,
            b.phone,
            b.created_at.strftime("%Y-%m-%d %H:%M") if b.created_at else "",
        ])
    _autosize(history)

    buffer = BytesIO()
    book.save(buffer)
    return buffer.getvalue()
