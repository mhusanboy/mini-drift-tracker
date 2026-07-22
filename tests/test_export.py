from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook

from bot.db.models import Booking, BookingStatus
from bot.services.export import build_stats_workbook
from bot.services.stats import UserStat

TODAY = date(2026, 7, 20)

OVERVIEW = {
    "users": 2, "requests": 4, "accepted": 2, "rejected": 1, "pending": 1,
    "today": 1, "people": 13, "hours": 3,
}
USERS = [
    UserStat(name="Anvar", phone="+998901234567", language="uz", requests=3,
             accepted=1, people=4, first_seen=date(2026, 7, 1), last_booking=TODAY),
    UserStat(name="Bek", phone="+998900000000", language="ru", requests=1,
             accepted=1, people=9, first_seen=None, last_booking=None),
]
BOOKINGS = [
    Booking(id=1, user_id=1, full_name="Anvar", phone="+998901234567",
            when_text="ertaga 18:00", date=TODAY, start_minute=18 * 60,
            people_count=4, duration_hours=1, status=BookingStatus.ACCEPTED,
            created_at=datetime(2026, 7, 19, 10, 30)),
    Booking(id=2, user_id=2, full_name="Bek", phone="+998900000000",
            when_text="shanba kechqurun", date=None, start_minute=None,
            people_count=9, duration_hours=2, status=BookingStatus.PENDING,
            created_at=datetime(2026, 7, 19, 11, 0)),
]


def _book():
    return load_workbook(BytesIO(build_stats_workbook(OVERVIEW, USERS, BOOKINGS, TODAY, "uz")))


def test_workbook_has_the_three_sheets():
    assert _book().sheetnames == ["Umumiy", "Mijozlar", "Bronlar"]


def test_overview_sheet_carries_every_metric():
    values = [row[1] for row in _book()["Umumiy"].iter_rows(values_only=True) if row[1] is not None]
    assert values == [2, 4, 2, 1, 1, 1, 13, 3]


def test_customer_rows():
    rows = list(_book()["Mijozlar"].iter_rows(min_row=2, values_only=True))
    assert rows[0][:6] == ("Anvar", "+998901234567", "uz", 3, 1, 4)
    # A user who never booked leaves the date cells empty, not the word "None".
    assert rows[1][6] is None and rows[1][7] is None


def test_booking_rows_show_the_resolved_span():
    rows = list(_book()["Bronlar"].iter_rows(min_row=2, values_only=True))
    assert rows[0][1] == "2026-07-20 18:00–19:00"
    assert rows[0][2] == "ertaga 18:00"   # what the customer actually typed


def test_an_unresolved_booking_falls_back_to_the_raw_text():
    rows = list(_book()["Bronlar"].iter_rows(min_row=2, values_only=True))
    assert rows[1][0] is None                  # no date to put in the date column
    assert rows[1][1] == "shanba kechqurun"    # the words instead of a span


def test_status_is_localized_not_raw():
    rows = list(_book()["Bronlar"].iter_rows(min_row=2, values_only=True))
    assert rows[0][5] == "✅ qabul qilindi"
    assert rows[1][5] == "⏳ kutilmoqda"
